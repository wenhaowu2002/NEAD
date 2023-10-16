import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# 打开原工作簿
workbook = openpyxl.load_workbook('raw_data.xlsx', data_only=True)

# 创建一个新的工作簿
new_workbook = openpyxl.Workbook()

# 需要排除的国家名称列表
excluded_countries = [
    "Dominica", "Falkland Is.", "Faroe Is.", "Fiji", "Gaza Strip", "Gibraltar", "Grenada",
    "Guadeloupe", "Guam", "Isle of Man", "Jan Mayen", "Kiribati", "Liechtenstein", "Maldives",
    "Malta", "Marshall Is.", "Martinique", "Mayotte", "Micronesia", "Monaco", "Montserrat",
    "Nauru", "Netherlands Antilles", "New Caledonia", "Niue", "Norfolk I.", "Northern Mariana Is.",
    "Palau", "Pitcairn Is.", "Samoa", "San Marino", "Sao Tome & Principe", "Seychelles",
    "Solomon Is.", "South Georgia & the South Sandwich Is.", "St. Helena", "St. Kitts & Nevis",
    "St. Lucia", "St. Pierre & Miquelon", "St. Vincent & the Grenadines", "Svalbard", "Tokelau",
    "Tonga", "Turks & Caicos Is.", "Tuvalu", "Vanuatu", "Virgin Is.", "Wallis & Futuna", "West Bank"
]

# 遍历每个工作表
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # 从数据框中创建一个新的工作表
    df = pd.DataFrame(sheet.values)

    # 筛选掉排除的国家
    filtered_df = df[~df[0].isin(excluded_countries)]

    # 创建一个新的工作表
    new_sheet = new_workbook.create_sheet(title=sheet_name)

    # 将筛选后的数据框写入新工作表
    for row in dataframe_to_rows(filtered_df, index=False, header=False):
        new_sheet.append(row)

# 删除新工作簿默认创建的工作表
new_workbook.remove(new_workbook.active)

# 保存新的工作簿
new_workbook.save('raw_data_.xlsx')
